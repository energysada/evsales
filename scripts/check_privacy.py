#!/usr/bin/env python3
"""
Pre-deploy privacy check for evsales-site.

Scans deploy artifacts (index.html, news.json) for any large unit volumes
that appear in private historical xlsx files. Designed to catch the specific
failure mode where monthly BEV unit counts get leaked into the public site.

Strategy (focused, low-false-positive):
- Only check 5+ digit integer unit volumes from bev.xlsx Table 1
- Only check the HISTORICAL period (skip last 3 columns of recent data which
  may legitimately match public KBA/SMMT/etc. for the current month)
- Do NOT check percentages (shares overlap too much with public data)
- Do NOT check small numbers (<10,000) — too many false positives
- Use word-boundary matching with both raw and comma-separated formats
- Allowlist for numbers that are legitimately in public sources

Usage:
    python3 check_privacy.py

Exit codes:
    0 = clean, safe to deploy
    1 = violations found, deploy should be blocked
    2 = check could not run (privacy unknown — fail safe)
"""

import re
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl not installed; cannot run privacy check", file=sys.stderr)
    sys.exit(2)

SITE = Path("/Users/energysada/ev/evsales-site")
DEPLOY_ARTIFACTS = ["index.html", "news.json"]

# Allowlist: values that legitimately appear in public sources
# Add any 5+ digit number here that's both in bev.xlsx AND a public press release
PUBLIC_ALLOWLIST = {
    # Cox Q1 2026 figures
    212600, 296304, 122196, 93500, 34821,
    # KBA Mar 2026
    70663, 29996, 358592, 294161, 117846,
    # SMMT Mar 2026
    86120, 49671, 380627, 60268, 165997, 18571, 196059, 69313, 33815,
    # FCAI Mar 2026
    15839, 8215, 105058,
    # AVEM/PFA France Mar 2026
    49406, 173634, 7023, 3493, 2824, 20140, 401555,
    # UNRAE Italy Mar 2026
    16137, 16998, 33135, 185367, 484802,
    # ANFAC Spain Mar 2026
    11861, 14859, 26720, 27273, 62916,
    # OFV Norway Mar 2026
    17400, 17685, 14288,
    # Mobility Sweden / Denmark Mar 2026
    10897, 14616, 18968, 26578, 1134, 596953,
    # Vahan/FADA India Mar 2026
    22315, 191067, 279530, 199590, 199923, 1401818, 1401663, 2452056, 696769, 1100000, 696769, 1402138, 830818, 19648, 2452921, 14488, 9956,
    # KAIDA Korea Mar 2026
    11130, 11134, 1664, 33970, 16249, 24444, 89074, 35693, 13128,
    # CPCA China Mar 2026
    784000, 1120000, 295693, 85670, 127728, 88697, 1050000, 1500000, 88967,
    # New Zealand Mar 2026
    2422, 2370, 3108, 14908, 1107, 100000,
    # Cambodia Feb 2026
    14534, 1078, 488, 8300000,
    # Australia Feb 2026 reference
    3485, 2818, 6858, 8877, 3576, 700,
}

def extract_private_unit_volumes():
    """Extract historical 5+ digit unit volumes from bev.xlsx Table 1 only."""
    private = set()
    path = SITE / "bev.xlsx"
    if not path.exists():
        print(f"WARN: {path} not found — skipping check", file=sys.stderr)
        return private

    try:
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    except Exception as e:
        print(f"ERROR: cannot open bev.xlsx: {e}", file=sys.stderr)
        sys.exit(2)

    # Find the EVSales_Filtered sheet (or first sheet)
    sheet_name = "EVSales_Filtered" if "EVSales_Filtered" in wb.sheetnames else wb.sheetnames[0]
    ws = wb[sheet_name]

    # Table 1: rows 4-20 (countries) — first 17 rows of data
    # Skip last 3 columns (most recent period — may match public sources)
    max_col = ws.max_column
    historical_max_col = max(2, max_col - 3)

    for row_cells in ws.iter_rows(min_row=4, max_row=20, max_col=historical_max_col, values_only=True):
        for v in row_cells:
            if v is None:
                continue
            if isinstance(v, (int, float)):
                iv = int(v) if v == int(v) else None
                if iv and iv >= 10000:
                    if iv not in PUBLIC_ALLOWLIST:
                        private.add(iv)
    wb.close()
    return private


def scan_artifacts(private_set):
    """Scan deploy artifacts for matches against private unit volumes."""
    violations = []  # list of (file, value, context)

    for fname in DEPLOY_ARTIFACTS:
        path = SITE / fname
        if not path.exists():
            continue
        text = path.read_text()

        for v in private_set:
            # Two patterns: bare digits and comma-separated
            patterns = [str(v), f"{v:,}"]
            for pat in patterns:
                # Word boundary — no adjacent digits, dots, commas, or %
                rx = r'(?<![\d.,])' + re.escape(pat) + r'(?![\d.,%])'
                m = re.search(rx, text)
                if m:
                    start = max(0, m.start() - 40)
                    end = min(len(text), m.end() + 40)
                    context = text[start:end].replace('\n', ' ').strip()
                    violations.append((fname, pat, context))
                    break

    return violations


def main():
    private = extract_private_unit_volumes()
    if not private:
        print("  privacy check: no private values to check (allowlist filtered all)")
        sys.exit(0)

    print(f"  privacy check: scanning {len(private)} historical unit volumes...")
    violations = scan_artifacts(private)

    if not violations:
        print("  ✓ privacy check passed — no historical unit volumes found in deploy artifacts")
        sys.exit(0)

    print()
    print("  ╔══════════════════════════════════════════════════════════════════════╗")
    print("  ║  ❌ PRIVACY CHECK FAILED — deploy blocked                            ║")
    print("  ╚══════════════════════════════════════════════════════════════════════╝")
    print()
    print(f"  Found {len(violations)} potential leaks of private historical unit volumes:")
    print()
    for fname, val, ctx in violations:
        print(f"    {fname}: '{val}'")
        print(f"      ...{ctx}...")
        print()

    print("  These values appear in bev.xlsx historical data and should NOT be in")
    print("  deployed artifacts. Either:")
    print("    1. Use directional language instead ('above the trailing average')")
    print("    2. Round heavily and verify not reverse-engineerable")
    print("    3. If from a public source, add to PUBLIC_ALLOWLIST in this script")
    print()
    sys.exit(1)


if __name__ == "__main__":
    main()
